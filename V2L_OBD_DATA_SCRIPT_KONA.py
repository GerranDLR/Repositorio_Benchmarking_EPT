import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook

def generar_graficos(ruta_entrada, ruta_salida):
    # Crear la carpeta de salida si no existe
    carpeta_graficos = os.path.join(ruta_salida, "graficos tratados")
    os.makedirs(carpeta_graficos, exist_ok=True)

    # Procesar cada archivo Excel
    for archivo in os.listdir(ruta_entrada):
        if archivo.endswith(".xlsx") or archivo.endswith(".xls"):
            ruta_archivo = os.path.join(ruta_entrada, archivo)
            try:
                wb = load_workbook(ruta_archivo, data_only=True)
                hoja = wb.active

                # Extraer y filtrar datos sincronizados
                datos = [
                    (
                        hoja.cell(row=i, column=4).value,  # Tiempo
                        hoja.cell(row=i, column=14).value, # Voltaje
                        hoja.cell(row=i, column=15).value, # Corriente
                        hoja.cell(row=i, column=11).value, # SOC
                        hoja.cell(row=i, column=12).value  # Potencia
                    )
                    for i in range(15, hoja.max_row + 1)
                ]
                datos = [fila for fila in datos if all(x is not None for x in fila)]

                tiempo = [x[0] for x in datos]
                columna_14 = [x[1] for x in datos]
                columna_15 = [x[2] for x in datos]
                columna_11 = [x[3] for x in datos]
                columna_12 = [x[4] for x in datos]

               # Gráfico 1: Voltaje, Corriente, SOC y Potencia
                fig, ax1 = plt.subplots(figsize=(10, 6))

                # Calcular la media de la potencia
                media_potencia = sum(columna_12) / len(columna_12)

                # Modificar la etiqueta del gráfico para incluir la media de la potencia
                ax1.plot(tiempo, columna_12, label=f"Power (kW) Min: {min(columna_12):.2f}, Avg: {media_potencia:.2f}", color="red")
                ax1.plot(tiempo, columna_15, label=f"HV Current (A) Min: {min(columna_15):.2f}", color="blue")
                ax1.set_xlabel("Time (s)")
                ax1.set_ylabel("Power (kW) / Current (A)", color="Black")
                ax1.tick_params(axis="y", labelcolor="Black")
                ax1.grid()

                ax2 = ax1.twinx()
                ax2.plot(tiempo, columna_11, label=f"SOC (%) Max: {max(columna_11):.2f}", color="black")
                ax2.plot(tiempo, columna_14, label=f"HV Voltage (V) Max: {max(columna_14):.2f}", color="green")
                ax2.set_ylabel("SOC (%) / HV Voltage", color="Black")
                ax2.tick_params(axis="y", labelcolor="Black")

                fig.suptitle("Graph 1: HV Voltage, HV Current, SOC & HV Power vs Time")

                # Ajustar la posición de las leyendas
                ax1.legend(loc="upper center", bbox_to_anchor=(0.5, -0.1), ncol=2)
                ax2.legend(loc="upper center", bbox_to_anchor=(0.5, -0.2), ncol=2)

                # Guardar el gráfico
                plt.savefig(os.path.join(carpeta_graficos, f"{archivo}_grafico_1.png"), bbox_inches="tight")
                plt.close()

                # Gráfico 3: Corriente vs Potencia
                plt.figure(figsize=(10, 6))
                plt.plot(columna_15, columna_12, label=f"Current vs Power Max: {max(columna_12):.2f}", color="blue")
                plt.title("Graph 3: Current vs Power")
                plt.xlabel("Power (kW)")
                plt.ylabel("Current (A)")
                plt.legend(loc="upper center", bbox_to_anchor=(0.5, -0.1), ncol=1)
                plt.grid()
                plt.savefig(os.path.join(carpeta_graficos, f"{archivo}_grafico_3.png"), bbox_inches="tight")
                plt.close()

                print(f"Gráficos generados y guardados para {archivo}")

            except Exception as e:
                print(f"Error al procesar el archivo {archivo}: {e}")

def generar_grafico_para_archivo_especifico(ruta_archivo, ruta_salida):
    # Crear la carpeta de salida si no existe
    carpeta_graficos = os.path.join(ruta_salida, "graficos tratados")
    os.makedirs(carpeta_graficos, exist_ok=True)

    try:
        wb = load_workbook(ruta_archivo, data_only=True)
        hoja = wb.active

        # Extraer y filtrar datos sincronizados
        datos = [
            (
                hoja.cell(row=i, column=4).value,  # Tiempo
                hoja.cell(row=i, column=14).value, # Voltaje
                hoja.cell(row=i, column=11).value, # SOC
                hoja.cell(row=i, column=25).value  # Temperatura
            )
            for i in range(15, hoja.max_row + 1)
        ]
        datos = [fila for fila in datos if all(x is not None for x in fila)]

        tiempo = [x[0] for x in datos]
        columna_14 = [x[1] for x in datos]
        columna_11 = [x[2] for x in datos]
        columna_25 = [x[3] for x in datos]

        # Gráfico: SOC y Temperatura en eje Y izquierdo, Voltaje en eje Y derecho
        fig, ax1 = plt.subplots(figsize=(10, 6))
        ax1.plot(tiempo, columna_11, label=f"SOC (%) Max: {max(columna_11):.2f}", color="black")
        ax1.plot(tiempo, columna_25, label=f"Temperature (°C) Max: {max(columna_25):.2f}", color="red")
        ax1.set_xlabel("Time (s)")
        ax1.set_ylabel("SOC (%) / Temperature (°C)", color="Black")
        ax1.tick_params(axis="y", labelcolor="Black")
        ax1.grid()

        ax2 = ax1.twinx()
        ax2.plot(tiempo, columna_14, label=f"Voltage (V) Max: {max(columna_14):.2f}", color="green")
        ax2.set_ylabel("Voltage (V)", color="Black")
        ax2.tick_params(axis="y", labelcolor="Black")

        fig.suptitle("Graph: 3h Endurance")
        ax1.legend(loc="upper center", bbox_to_anchor=(0.5, -0.1), ncol=2)
        ax2.legend(loc="upper center", bbox_to_anchor=(0.5, -0.2), ncol=1)

        nombre_grafico = os.path.join(carpeta_graficos, f"{os.path.basename(ruta_archivo)}_grafico_col25_11_14.png")
        plt.savefig(nombre_grafico, bbox_inches="tight")
        plt.close()

        print(f"Gráfico generado y guardado: {nombre_grafico}")

    except Exception as e:
        print(f"Error al procesar el archivo {ruta_archivo}: {e}")

# Rutas de entrada y salida
ruta_entrada = r"C:\Users\SNX6774\OneDrive - Nissan Motor Corporation\Escritorio\OBD tratado\Datos_OBD_Excel_CT"
ruta_salida = r"C:\Users\SNX6774\OneDrive - Nissan Motor Corporation\Escritorio\OBD tratado"

# Generar gráficos para todos los archivos
generar_graficos(ruta_entrada, ruta_salida)

# Generar gráfico específico para el archivo solicitado
ruta_archivo_especifico = r"C:\Users\SNX6774\OneDrive - Nissan Motor Corporation\Escritorio\OBD tratado\Datos_OBD_Excel_CT\20F4-10-31 09-37-13- 3h test.xlsx"
generar_grafico_para_archivo_especifico(ruta_archivo_especifico, ruta_salida)
