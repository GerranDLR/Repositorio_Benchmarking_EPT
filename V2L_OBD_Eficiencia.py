import os  
import matplotlib.pyplot as plt
from openpyxl import load_workbook

def calcular_y_graficar_eficiencia(ruta_entrada, ruta_salida):
    # Crear carpeta para gráficos
    carpeta_graficos = os.path.join(ruta_salida, "graficos tratados")
    os.makedirs(carpeta_graficos, exist_ok=True)

    # Mapeo de tests y sus potencias nominales
    pruebas = {
        "- eff 100W": 0.1,
        "- eff 500W - 20": 0.5,
        "- eff 500W": 0.5,
        "- eff 1000W": 1,
        "- eff 2000W": 2,
        "- eff 3000W": 3,
        "- eff 3300W": 3.3,
    }

    # Valores a restar en el denominador (correcciones)
    correcciones = {
        0.1: 0.23874 / 0.95,
        0.5: 0.23877 / 0.95,  # 500W
        0.6: 0.15857 / 0.95,  # 500W - 20
        1: 0.23878 / 0.95,
        2: 0.23859 / 0.95,
        3: 0.2384 / 0.95,
        3.3: 0.2389 / 0.95,
    }

    # Almacenar potencias medias y eficiencias
    resultados = {test: None for test in pruebas}

    # Procesar cada archivo Excel
    for archivo in os.listdir(ruta_entrada):
        if archivo.endswith(".xlsx") or archivo.endswith(".xls"):
            ruta_archivo = os.path.join(ruta_entrada, archivo)
            try:
                test_asociado = next((test for test in pruebas if test in archivo), None)
                if not test_asociado:
                    print(f"Archivo {archivo} no corresponde a ningún test definido. Saltando.")
                    continue

                wb = load_workbook(ruta_archivo, data_only=True)
                hoja = wb.active

                # Extraer datos (tiempo y potencia)
                datos = [
                    (
                        hoja.cell(row=i, column=4).value,  # Tiempo
                        hoja.cell(row=i, column=12).value  # Potencia
                    )
                    for i in range(15, hoja.max_row + 1)
                ]
                datos = [fila for fila in datos if all(x is not None for x in fila)]  # Filtrar datos válidos

                # Filtrar datos en el rango de tiempo [25, 170] segundos
                datos_filtrados = [p for t, p in datos if 25 <= t <= 170]

                if not datos_filtrados:
                    print(f"No hay datos en el rango temporal para {archivo}. Saltando archivo.")
                    continue

                # Calcular potencia media y convertir a watts
                potencia_media = sum(datos_filtrados) / len(datos_filtrados)  # Potencia media en kW
                resultados[test_asociado] = abs(potencia_media)

                print(f"Archivo: {archivo} | Test: {test_asociado} | Potencia media: {potencia_media:.2f} kW")

            except Exception as e:
                print(f"Error al procesar el archivo {archivo}: {e}")

    # Preparar datos para graficar
    etiquetas = []
    eficiencias = []
    eficiencias_corregidas = []
    dot_label = "500W"
    dot_eficiencia_azul = None
    dot_eficiencia_rojo = None

    for test, potencia_nominal in pruebas.items():
        potencia_media = resultados[test]
        if potencia_media is not None:
            # Caso especial para "- eff 500W - 20"
            if test == "- eff 500W - 20":
                # Eficiencia sin corrección (azul), usando 500W (0.5) como potencia nominal
                eficiencia = (0.5 / potencia_media) * 100
                dot_eficiencia_azul = eficiencia

                # Corrección específica asociada a 0.6
                correccion = correcciones.get(0.6, 0)
                if potencia_media - correccion > 0:
                    # Eficiencia corregida (roja), también usando 500W (0.5)
                    dot_eficiencia_rojo = (0.5 / (potencia_media - correccion)) * 100
                else:
                    dot_eficiencia_rojo = 0

                continue  # Ignorar este punto en las curvas principales

            # Para otros tests
            eficiencia = (potencia_nominal / potencia_media) * 100
            etiquetas.append(test.replace("- eff ", ""))
            eficiencias.append(eficiencia)

            correccion = correcciones.get(potencia_nominal, 0)
            if potencia_media - correccion > 0:
                eficiencia_corregida = (potencia_nominal / (potencia_media - correccion)) * 100
            else:
                eficiencia_corregida = 0
            eficiencias_corregidas.append(eficiencia_corregida)

            print(f"Test: {test} | Potencia nominal: {potencia_nominal} kW | "
                  f"Eficiencia: {eficiencia:.2f}% | Eficiencia corregida: {eficiencia_corregida:.2f}%")

    if etiquetas and eficiencias:
        # Graficar
        plt.figure(figsize=(10, 6))
        plt.plot(etiquetas, eficiencias, marker='o', linestyle='-', color='b', label="System efficiency")
        plt.plot(etiquetas, eficiencias_corregidas, marker='o', linestyle='--', color='r', label="Component efficiency")

        # Agregar puntos aislados
        if dot_eficiencia_azul is not None:
            plt.scatter([dot_label], [dot_eficiencia_azul], color='b', label="500W-20 Point (System)", zorder=5)
        if dot_eficiencia_rojo is not None:
            plt.scatter([dot_label], [dot_eficiencia_rojo], color='r', label="500W-20 Point (Component)", zorder=5)

        # Agregar flechas y etiquetas
        if dot_eficiencia_azul and dot_eficiencia_rojo:
            plt.annotate(
                "500W_20% SOC",
                xy=(dot_label, dot_eficiencia_azul),
                xytext=(dot_label, dot_eficiencia_azul - 5),
                fontsize=10,
                color='blue'
            )
            plt.annotate(
                "500W_20% SOC",
                xy=(dot_label, dot_eficiencia_rojo),
                xytext=(dot_label, dot_eficiencia_rojo + 4),
                fontsize=10,
                color='red'
            )

        # Título y etiquetas
        plt.title("Efficiency Curve")
        plt.xlabel("Test")
        plt.ylabel("Efficiency (%)")
        plt.ylim(0, 100)
        plt.grid(True, linestyle='--', alpha=0.7)

        # Colocar la leyenda debajo del gráfico
        plt.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=2)

        plt.tight_layout()

        # Guardar gráfico
        grafico_path = os.path.join(carpeta_graficos, "curva_eficiencia.png")
        plt.savefig(grafico_path, dpi=150)
        plt.close()

        print(f"Gráfico guardado en {grafico_path}.")


# Rutas de entrada y salida
ruta_entrada = r"C:\Users\SNX6774\OneDrive - Nissan Motor Corporation\Escritorio\OBD tratado\Datos_OBD_Excel_CT"
ruta_salida = r"C:\Users\SNX6774\OneDrive - Nissan Motor Corporation\Escritorio\OBD tratado"

# Llamar a la función
calcular_y_graficar_eficiencia(ruta_entrada, ruta_salida)
